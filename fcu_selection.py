# -*- coding: utf-8 -*-
"""
  ρ_da = 1 / GetMoistAirVolume(tW, dW)   [室外干空气密度, kg_da/m³]
  cp_moist = 1.006 + 1.86×dW              [湿空气质量比热, kJ/kg·°C]
  rho_cp = ρ_da × cp_moist                [有效体积比热, kJ/(m³·°C)]

  Vs = Q_room × 3600 / (ρ_da × Δh_SHF)   [SHF线焓差法]
  Q_rise = Vf × rho_cp × dT_rise / 3600
  Q_ahu  = Vf × ρ_ahu × (hW - hL) / 3600  其中 ρ_ahu 由后续说明
  tL     = 二分法 h(tL, RH_L=90%) = hN    [室内等焓]
  Q_fcu_sen 所用密度 = ρ_da × 1.006
"""

import itertools
import warnings
from collections import Counter

import openpyxl
import pandas as pd
import psychrolib
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


psychrolib.SetUnitSystem(psychrolib.SI)

FCU_DB = {
    "FP-34":  {"airflow": 340,  "cooling": 1800,  "resistance": 30},
    "FP-51":  {"airflow": 510,  "cooling": 2700,  "resistance": 30},
    "FP-68":  {"airflow": 680,  "cooling": 3600,  "resistance": 30},
    "FP-85":  {"airflow": 850,  "cooling": 4500,  "resistance": 30},
    "FP-102": {"airflow": 1020, "cooling": 5400,  "resistance": 40},
    "FP-136": {"airflow": 1360, "cooling": 7200,  "resistance": 40},
    "FP-170": {"airflow": 1700, "cooling": 9000,  "resistance": 40},
    "FP-204": {"airflow": 2040, "cooling": 10800, "resistance": 40},
    "FP-238": {"airflow": 2380, "cooling": 12600, "resistance": 50},
}
FCU_MODELS = list(FCU_DB.keys())


INPUT_FILE = "计算书.xlsx"
SHEET_NAME = "Sheet1"


def safe_float(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def load_data(filepath, sheet):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet]
    rows = []
    current_floor = None
    for r in ws.iter_rows(min_row=1, values_only=True):
        if r[0] is not None and isinstance(r[0], str):
            v = r[0].strip()
            if "层" in v:
                current_floor = v
            if r[1] is None:
                continue
        room = r[1]
        if room is None:
            continue
        cooling = safe_float(r[3])
        humid   = safe_float(r[4])
        if cooling is None or cooling <= 0 or humid is None:
            continue
        fresh_vol = safe_float(r[5])
        if fresh_vol is None:
            continue
        rows.append({
            "楼层":           current_floor,
            "房间":           str(room).strip(),
            "面积m2":         safe_float(r[2]),
            "冷负荷kW":       cooling,
            "湿负荷kg/h":     humid,
            "新风量m3/h":     fresh_vol,
            "送风温差°C":     safe_float(r[6]) or 8.0,
            "新风机露点RH%":  safe_float(r[7]) or 90.0,
            "新风管道温升°C": safe_float(r[8]) or 2.0,
            "室内干球°C":     safe_float(r[9]) or 25.0,
            "室内相对湿度%":  safe_float(r[10]) or 60.0,
            "室外干球°C":     safe_float(r[11]) or 34.7,
            "室外相对湿度%":  safe_float(r[12]) or 54.6,
        })
    return pd.DataFrame(rows)



def calc_row(row: dict) -> dict:
    Q       = row["冷负荷kW"]
    W       = row["湿负荷kg/h"]
    Vf      = row["新风量m3/h"]
    dT      = row["送风温差°C"]
    rh_L    = row["新风机露点RH%"] / 100.0
    dT_rise = row["新风管道温升°C"]
    tN      = row["室内干球°C"]
    rhN     = row["室内相对湿度%"] / 100.0
    tW      = row["室外干球°C"]
    rhW     = row["室外相对湿度%"] / 100.0

    #   室内状态点 (N) 
    dN = psychrolib.GetHumRatioFromRelHum(tN, rhN, 101325)
    hN = psychrolib.GetMoistAirEnthalpy(tN, dN) / 1000.0      # kJ/kg_da

    # 室外状态点 (W) 
    dW    = psychrolib.GetHumRatioFromRelHum(tW, rhW, 101325)
    hW    = psychrolib.GetMoistAirEnthalpy(tW, dW) / 1000.0   # kJ/kg_da
    rhoW  = psychrolib.GetMoistAirDensity(tW, dW, 101325)     # kg_ma/m³（湿空气密度）
    vW    = psychrolib.GetMoistAirVolume(tW, dW, 101325)       # m³/kg_da
    rho_da = 1.0 / vW                                          # kg_da/m³（干空气密度）

    # 湿空气比热：kJ/(kg_da·°C)
    cp_moist = 1.006 + 1.86 * dW
    # 有效体积比热 [kJ/(m³·°C)]
    rho_cp = rho_da * cp_moist

    #  新风机出风点 (L)：h_L = h_N，RH_L 给定 
    def h_at(t, rh):
        d = psychrolib.GetHumRatioFromRelHum(t, rh, 101325)
        return psychrolib.GetMoistAirEnthalpy(t, d) / 1000.0

    lo, hi = 0.0, tN
    for _ in range(80):
        mid = (lo + hi) / 2
        if h_at(mid, rh_L) < hN:
            lo = mid
        else:
            hi = mid
    tL = (lo + hi) / 2
    dL = psychrolib.GetHumRatioFromRelHum(tL, rh_L, 101325)
    hL = h_at(tL, rh_L)

    # 新风入室温度（经管道温升后）
    tIn = tL + dT_rise

    #  热湿比 & 送风状态 
    epsilon = (Q * 3600.0) / W if W > 0 else 0.0

    ts = tN - dT                               # 送风温度
    # 沿热湿比线求送风含湿量：h_s = hN - ε*(dN-ds), h_s = f(ts, ds)
    # → ds = (hN - ε*dN - 1.006*ts) / (2501+1.86*ts - ε)
    denom_shf = (2501 + 1.86 * ts) - epsilon
    if abs(denom_shf) > 1e-6:
        ds = (hN - epsilon * dN - 1.006 * ts) / denom_shf
    else:
        ds = dN
    hs = 1.006 * ts + ds * (2501 + 1.86 * ts)
    dh_shf = hN - hs                           # kJ/kg_da

    #  送风量 & 回风量 
    # 天正公式：Vs = Q × 3600 / (ρ_da × Δh_SHF)
    # 物理意义：ρ_da [kg_da/m³] × Vs [m³/h] / 3600 × Δh [kJ/kg_da] = Q [kW]
    if abs(dh_shf) > 1e-6:
        Vs = Q * 3600.0 / (rho_da * dh_shf)
    else:
        # Fallback to empirical formula if enthalpy difference is ~0
        Vs = (Q * 1000) / (0.35 * dT)
    
    Vr = Vs - Vf
    fresh_ratio = (Vf / Vs) * 100.0 if Vs > 0 else 0.0

    #  各设备冷负荷 
    # 新风管道温升负荷（显热）
    # rho_cp = ρ_da × cp_moist  [kJ/(m³·°C)]
    Q_rise = Vf * rho_cp * dT_rise / 3600.0

    # FCU 总冷量
    Q_fcu = Q + Q_rise

    # 新风 AHU 冷量：送入室内的新风焓差（以 kg_da 为基准）
    # Q = ρ_da × Vf/3600 × (hW - hL)
    Q_ahu = Vf * rho_da * (hW - hL) / 3600.0

    # FCU 显热冷量
    Q_room_sen = Q - W * 2500.0 / 3600.0
    # 新风显热冷量贡献：用 rho_cp（显热传热因子）
    Q_oa_sen   = Vf * rho_cp * (tN - tIn) / 3600.0
    Q_fcu_sen  = Q_room_sen - Q_oa_sen

    return {
        # 室内
        "hN(kJ/kg)":      round(hN, 3),
        "dN(kg/kg)":      round(dN, 6),
        # 室外
        "hW(kJ/kg)":      round(hW, 3),
        "dW(kg/kg)":      round(dW, 6),
        "ρW(kg/m³)":      round(rhoW, 4),
        "ρda(kg_da/m³)":  round(rho_da, 4),
        "vW(m³/kg_da)":   round(vW, 5),
        "cp_moist(kJ/kg)":round(cp_moist, 5),
        "ρ_cp(kJ/m³°C)":  round(rho_cp, 5),
        # 新风机出风
        "tL(°C)":         round(tL, 3),
        "dL(kg/kg)":      round(dL, 6),
        "hL(kJ/kg)":      round(hL, 3),
        "tIn(°C)":        round(tIn, 3),
        # 送风状态
        "ts(°C)":         round(ts, 1),
        "ds(kg/kg)":      round(ds, 6),
        "hs(kJ/kg)":      round(hs, 3),
        "Δh_SHF(kJ/kg)":  round(dh_shf, 3),
        # 风量
        "送风量Vs(m³/h)": round(Vs, 1),
        "回风量Vr(m³/h)": round(Vr, 1),
        "新风比(%)":       round(fresh_ratio, 3),
        "热湿比ε(kJ/kg)": round(epsilon, 2),
        # 负荷
        "Q_rise(kW)":     round(Q_rise, 5),
        "FCU冷量Q_fcu(kW)":       round(Q_fcu, 5),
        "AHU冷量Q_ahu(kW)":       round(Q_ahu, 5),
        "室内显热Q_room_sen(kW)": round(Q_room_sen, 5),
        "Q_oa_sen(kW)":           round(Q_oa_sen, 5),
        "FCU显热Q_fcu_sen(kW)":   round(Q_fcu_sen, 5),
    }


#  组合选型算法

def format_combo(combo):
    cnt = Counter(combo)
    return " + ".join(
        f"{cnt[m]}×{m}" for m in FCU_MODELS if m in cnt
    )


def find_fcu_combinations(Q_fcu_W: float, Vr: float, top_n: int = 3):
    candidates = []
    for n in range(1, 4):
        for combo in itertools.combinations_with_replacement(FCU_MODELS, n):
            tc = sum(FCU_DB[m]["cooling"] for m in combo)
            ta = sum(FCU_DB[m]["airflow"] for m in combo)
            if tc >= Q_fcu_W and ta >= Vr:
                max_a = max(FCU_DB[m]["airflow"] for m in combo)
                candidates.append({
                    "组合":           format_combo(combo),
                    "台数":           n,
                    "总风量(m³/h)":   ta,
                    "总制冷量(W)":    tc,
                    "最大单台风量":   max_a,
                    "风量余量(m³/h)": round(ta - Vr, 1),
                    "制冷量余量(W)":  round(tc - Q_fcu_W, 1),
                })
    candidates.sort(key=lambda x: (x["最大单台风量"], x["风量余量(m³/h)"]))
    return candidates[:top_n]


#  主程序
def main():
    print("读取...")
    df = load_data(INPUT_FILE, SHEET_NAME)
    print(f"   共读取 {len(df)} 个有效房间")

    print("执行计算与选型...")
    calc_list = []
    sel_list  = []

    for _, row in df.iterrows():
        calc = calc_row(row.to_dict())
        calc_list.append(calc)

        Q_fcu_W = calc["FCU冷量Q_fcu(kW)"] * 1000.0
        Vr      = calc["回风量Vr(m³/h)"]
        combos  = find_fcu_combinations(Q_fcu_W, max(Vr, 0), top_n=3)

        sel_list.append({
            "最优静音方案": combos[0]["组合"] if len(combos) > 0 else "无可行方案",
            "备选方案1":    combos[1]["组合"] if len(combos) > 1 else "—",
            "备选方案2":    combos[2]["组合"] if len(combos) > 2 else "—",
        })

    df_base = df.reset_index(drop=True)
    df_calc = pd.DataFrame(calc_list)
    df_sel  = pd.DataFrame(sel_list)

    #  选型汇总表
    df_selection = pd.concat([
        df_base[["楼层", "房间", "面积m2", "冷负荷kW", "湿负荷kg/h", "新风量m3/h"]],
        df_calc[["热湿比ε(kJ/kg)", "送风量Vs(m³/h)", "回风量Vr(m³/h)",
                 "FCU冷量Q_fcu(kW)", "AHU冷量Q_ahu(kW)", "FCU显热Q_fcu_sen(kW)",
                 "Q_rise(kW)"]].reset_index(drop=True),
        df_sel,
    ], axis=1)

    #  详细计算表
    df_detail = pd.concat([df_base, df_calc], axis=1)

    write_selection_excel(df_selection, "FCU_选型结果.xlsx")
    write_detail_excel(df_detail,       "FCU_详细计算.xlsx")

    print("   FCU_选型结果.xlsx  — 选型方案汇总")
    print("   FCU_详细计算.xlsx  — 全量计算过程")



# Excel 输出格式

THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 选型表 列组颜色
SEL_GROUPS = {
    "房间信息": (["楼层", "房间", "面积m2", "冷负荷kW", "湿负荷kg/h", "新风量m3/h"],
                 "D6E4F0"),
    "计算结果": (["热湿比ε(kJ/kg)", "送风量Vs(m³/h)", "回风量Vr(m³/h)",
                  "FCU冷量Q_fcu(kW)", "AHU冷量Q_ahu(kW)", "FCU显热Q_fcu_sen(kW)",
                  "Q_rise(kW)"],
                 "D5E8D4"),
    "选型方案": (["最优静音方案", "备选方案1", "备选方案2"],
                 "FFE6CC"),
}

# 三方案的表头颜色
SCHEME_HDR = {"最优静音方案": "70AD47", "备选方案1": "5B9BD5", "备选方案2": "ED7D31"}
SCHEME_ROW = {"最优静音方案": "EBF3E4", "备选方案1": "DAE9F5", "备选方案2": "FAE5D3"}


def _col_color(col_name, groups):
    for _gname, (cols, color) in groups.items():
        if col_name in cols:
            return color
    return "F5F5F5"


def _write_header(ws, headers, groups, row=1):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row, ci, h)
        c.fill   = PatternFill("solid", fgColor=_col_color(h, groups))
        c.font   = Font(bold=True, size=10, color="1F2D3D")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 36


def _write_data(ws, df, groups, scheme_cols=None, start_row=2):
    for ri, (_, row_data) in enumerate(df.iterrows()):
        ro = start_row + ri
        fill_color = "F9F9F9" if ri % 2 == 0 else "FFFFFF"
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(ro, ci)
            val = row_data[col]
            c.value = val if not (isinstance(val, float) and
                                  __import__("math").isnan(val)) else ""
            c.border    = BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font      = Font(size=10)
            # 方案列底色
            if scheme_cols and col in scheme_cols:
                c.fill = PatternFill("solid", fgColor=scheme_cols[col])
            else:
                c.fill = PatternFill("solid", fgColor=fill_color)


def _auto_width(ws, min_w=9, max_w=45):
    for col in ws.columns:
        ltr = get_column_letter(col[0].column)
        mx = 0
        for cell in col:
            s = str(cell.value) if cell.value is not None else ""
            w = sum(2 if ord(ch) > 127 else 1 for ch in s)
            mx = max(mx, w)
        ws.column_dimensions[ltr].width = min(max(mx + 2, min_w), max_w)


def write_selection_excel(df: pd.DataFrame, filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FCU选型汇总"

    headers = list(df.columns)
    _write_header(ws, headers, SEL_GROUPS)

    # 方案列表头颜色
    for ci, h in enumerate(headers, 1):
        if h in SCHEME_HDR:
            c = ws.cell(1, ci)
            c.fill = PatternFill("solid", fgColor=SCHEME_HDR[h])
            c.font = Font(bold=True, size=10, color="FFFFFF")

    # 数据行方案列颜色映射
    scheme_map = {k: v for k, v in SCHEME_ROW.items()}
    _write_data(ws, df, SEL_GROUPS, scheme_cols=scheme_map)

    # 方案列加宽
    for ci, h in enumerate(headers, 1):
        if h in SCHEME_HDR:
            ws.column_dimensions[get_column_letter(ci)].width = 40
    # 其余列自动
    _auto_width(ws)
    ws.freeze_panes = "B2"

    # 添加说明 sheet
    ws2 = wb.create_sheet("公式说明")
    notes = [
        ("参数", "公式 / 说明"),
        ("ρ_da (kg_da/m³)", "1 / GetMoistAirVolume(tW, dW, 101325)  — 室外干空气密度"),
        ("cp_moist (kJ/kg·°C)", "1.006 + 1.86 × dW  — 湿空气质量比热"),
        ("ρ_cp (kJ/m³·°C)", "ρ_da × cp_moist  — 体积比热（天正统一用此值）"),
        ("Δh_SHF (kJ/kg_da)", "hN − hs，hs 由热湿比线在 ts=tN−dT 处求出"),
        ("送风量 Vs (m³/h)", "Q × 3600 / (ρ_da × Δh_SHF)"),
        ("回风量 Vr (m³/h)", "Vs − Vf"),
        ("热湿比 ε (kJ/kg)", "Q × 3600 / 湿负荷"),
        ("tL (°C)", "二分法求 h(tL, RH_L) = hN（新风机出风与室内等焓）"),
        ("Q_rise (kW)", "Vf × ρ_cp × 管道温升 / 3600"),
        ("Q_fcu (kW)", "冷负荷 + Q_rise"),
        ("Q_ahu (kW)", "Vf × ρ_da × (hW − hL) / 3600"),
        ("Q_room_sen (kW)", "冷负荷 − 湿负荷 × 2500 / 3600"),
        ("Q_oa_sen (kW)", "Vf × ρ_cp × (tN − tIn) / 3600，tIn = tL + 管道温升"),
        ("Q_fcu_sen (kW)", "Q_room_sen − Q_oa_sen"),
    ]
    for row_data in notes:
        ws2.append(list(row_data))
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 65
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)

    wb.save(filename)
    print(f"   ✔ {filename}")


# 详细计算表列组
DETAIL_GROUPS = {
    "输入参数":  (["楼层","房间","面积m2","冷负荷kW","湿负荷kg/h",
                   "新风量m3/h","送风温差°C","新风机露点RH%","新风管道温升°C",
                   "室内干球°C","室内相对湿度%","室外干球°C","室外相对湿度%"],
                  "DDEEFF"),
    "室外参数":  (["hW(kJ/kg)","dW(kg/kg)","ρW(kg/m³)","ρda(kg_da/m³)",
                   "vW(m³/kg_da)","cp_moist(kJ/kg)","ρ_cp(kJ/m³°C)"],
                  "E2EFDA"),
    "室内参数":  (["hN(kJ/kg)","dN(kg/kg)"],
                  "FFF2CC"),
    "新风机出风": (["tL(°C)","dL(kg/kg)","hL(kJ/kg)","tIn(°C)"],
                   "FCE4D6"),
    "送风状态":  (["ts(°C)","ds(kg/kg)","hs(kJ/kg)","Δh_SHF(kJ/kg)"],
                  "EAD1DC"),
    "风量":      (["送风量Vs(m³/h)","回风量Vr(m³/h)","新风比(%)","热湿比ε(kJ/kg)"],
                  "D5E8D4"),
    "负荷":      (["Q_rise(kW)","FCU冷量Q_fcu(kW)","AHU冷量Q_ahu(kW)",
                   "室内显热Q_room_sen(kW)","Q_oa_sen(kW)","FCU显热Q_fcu_sen(kW)"],
                  "D6E4F0"),
}


def write_detail_excel(df: pd.DataFrame, filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "详细计算"

    headers = list(df.columns)
    _write_header(ws, headers, DETAIL_GROUPS)
    _write_data(ws, df, DETAIL_GROUPS)
    _auto_width(ws)
    ws.freeze_panes = "C2"

    wb.save(filename)
    print(f"   {filename}")


if __name__ == "__main__":
    main()
